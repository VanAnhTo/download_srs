import os
import time
import datetime

import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import DownloadSRS.my_config_downloadSRS


def firefoxProfile():
    profile = webdriver.FirefoxProfile()
    profile.set_preference('browser.download.folderList', 2)
    profile.set_preference('browser.download.manager.showWhenStarting', False)
    # profile.set_preference('browser.download.dir', os.getcwd())
    profile.set_preference('browser.download.dir', DownloadSRS.my_config_downloadSRS.raw_path_to_save)
    profile.set_preference('browser.helperApps.neverAsk.saveToDisk',
                           'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel')
    return profile

logging.basicConfig(filename='logFile.txt',
                    level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')
logging.debug('Start of program')

srs_id_list_wb = openpyxl.load_workbook(DownloadSRS.my_config_downloadSRS.srs_list_file)
sheet_srs_id = srs_id_list_wb[DownloadSRS.my_config_downloadSRS.sheet_name_srs]

numberOfId = sheet_srs_id.max_row

driver = webdriver.Firefox(executable_path=DownloadSRS.my_config_downloadSRS.firefox_geckodriver, firefox_profile = firefoxProfile())
driver.maximize_window()

driver.get('http:xxxx')
time.sleep(2)

username = driver.find_element_by_id("user")
# noinspection PyInterpreter
username.send_keys(DownloadSRS.my_config_downloadSRS.username)

pas = driver.find_element_by_id("password")
pas.send_keys(DownloadSRS.my_config_downloadSRS.password)

btnLogin = driver.find_element_by_css_selector("input.login_button")
btnLogin.click()

time.sleep(2)
#for r in range(2, 3):
for r in range (2, numberOfId+1):
    id_value = sheet_srs_id.cell(row=r, column=2).value
    logging.debug('ID value: ' + str(id_value))
    driver.get('http://xxxx/cb/exportRequirementsAsDocx.spr?tracker_id=' + str(id_value) + '&&viewId=-2')

    btnExportToExcelTab = driver.find_element_by_id("excelExportTabPane-tab")
    btnExportToExcelTab.click()

    radioRoundTripSelected = driver.find_element_by_id("roundtripExcelExport")
    radioRoundTripSelected.click()

    checkboxExportDescription = driver.find_element_by_css_selector("div.selectedExportBlock input#addDescriptionToExcelExportCheckbox")
    checkboxExportDescription.click()

    btbExport = driver.find_element_by_css_selector("div#excelExportTabPane input:nth-child(1).button")
    driver.execute_script("var elem=arguments[0]; setTimeout(function() {elem.click();}, 100)", btbExport)
    time.sleep(8)

driver.quit()


